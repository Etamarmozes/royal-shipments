#!/usr/bin/env python
"""Post-sync cleanup — surgical removal of email_updates / email_attachments /
shipment_events that came in from an unintended Gmail sync after the
full operational reset.

Usage:
    python backend/scripts/post_sync_cleanup.py --confirm POST_SYNC_CLEANUP

Pre-conditions checked at runtime:
    - shipments == 0   (we never want to delete email_updates that are tied
                        to live shipments — this script is for the specific
                        post-reset accidental-sync case only)
    - containers == 0
    - import_batches == 0

What gets cleared (operational data inserted after the reset):
    - email_attachments   (all rows; archived flag ignored — there are no
                           legit linked rows because shipments==0)
    - email_updates       (all rows)
    - shipment_events     (all rows; we know they are gmail_fetched +
                           classification rows from the unintended sync,
                           never linked to a real shipment because no
                           shipment exists)
    - pending_document_updates, pending_shipments, pending_containers,
      alerts, document_assignment_qc_results, document_assignment_actions
      → cleared if non-zero (defensive — should already be 0)

What is preserved (NOT touched):
    - users + roles + password hashes
    - document_assignment_rules (config / supplier rules)
    - shipments / containers / import_batches  (already 0; assertion failure
      stops the script if any are non-zero, so we never delete operational
      data the user wants to keep)
    - All code, configuration, env vars, OAuth credentials
    - The previous reset archive at uploads/archive_before_reset/<earlier-ts>/

Outputs (timestamped <ts> = %Y%m%d_%H%M%S):
    - DB backup           : backend/data/backups/royal_linen_before_post_sync_cleanup_<ts>.db
    - Audit xlsx          : backend/exports/post_sync_cleanup_audit_<ts>.xlsx
    - Log file            : backend/logs/post_sync_cleanup_<ts>.log
    - File archive folder : backend/uploads/archive_before_reset/<ts>/
"""
from __future__ import annotations

import argparse
import json
import logging
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime
from pathlib import Path

CONFIRMATION_TOKEN = "POST_SYNC_CLEANUP"

BACKEND_ROOT = Path(__file__).resolve().parent.parent
DB_PATH      = BACKEND_ROOT / "data" / "royal_linen.db"
BACKUPS_DIR  = BACKEND_ROOT / "data" / "backups"
EXPORTS_DIR  = BACKEND_ROOT / "exports"
LOGS_DIR     = BACKEND_ROOT / "logs"
UPLOADS_DOCS = BACKEND_ROOT / "uploads" / "documents"
ARCHIVE_ROOT = BACKEND_ROOT / "uploads" / "archive_before_reset"
HEALTH_SCRIPT = BACKEND_ROOT / "scripts" / "check_app_health.py"

# Tables we delete from. Order matters: children first.
SCOPED_TABLES_DELETE_ORDER = [
    "alerts",
    "document_assignment_actions",
    "document_assignment_qc_results",
    "pending_document_updates",
    "shipment_events",
    "email_attachments",
    "pending_containers",
    "pending_shipments",
    "email_updates",
]

# Tables we will REFUSE to touch — and assert are at zero before we run.
# If any of these is non-zero, this script aborts.  Use full_operational_reset.py
# for full clearing.
REQUIRE_EMPTY_BEFORE = [
    "shipments",
    "containers",
    "import_batches",
    "extra_work_tasks",
]

# Preserved tables (just for reporting).
PRESERVED_TABLES = [
    "users",
    "document_assignment_rules",
]


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument(
        "--confirm",
        default="",
        help=f"Must equal '{CONFIRMATION_TOKEN}' to proceed.",
    )
    return ap.parse_args()


def setup_logging(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("post_sync_cleanup")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    sh = logging.StreamHandler()
    sh.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, bytes):
        try:
            return v.decode("utf-8", errors="replace")
        except Exception:
            return repr(v)
    try:
        return json.dumps(v, ensure_ascii=False)
    except Exception:
        return str(v)


# ----- step 1: pre-flight assertions -----

def step_preflight(log: logging.Logger):
    log.info("STEP 1: pre-flight invariant check")
    if not DB_PATH.exists():
        raise RuntimeError(f"DB not found: {DB_PATH}")

    con = sqlite3.connect(str(DB_PATH))
    try:
        cur = con.cursor()
        for t in REQUIRE_EMPTY_BEFORE:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError as e:
                raise RuntimeError(f"required-empty table {t} missing: {e}")
            log.info("  %-30s = %d (required: 0)", t, n)
            if n != 0:
                raise RuntimeError(
                    f"REFUSED — table '{t}' has {n} rows. This script only operates "
                    f"on the post-reset accidental-sync state. Use "
                    f"full_operational_reset.py for full clearing."
                )
    finally:
        con.close()


# ----- step 2: backup DB -----

def step_backup_db(log: logging.Logger, ts: str) -> Path:
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    out = BACKUPS_DIR / f"royal_linen_before_post_sync_cleanup_{ts}.db"
    log.info("STEP 2: backing up DB -> %s", out)
    src = sqlite3.connect(str(DB_PATH))
    try:
        dst = sqlite3.connect(str(out))
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()
    if not out.exists() or out.stat().st_size == 0:
        raise RuntimeError(f"backup file missing or empty: {out}")
    chk = sqlite3.connect(str(out))
    try:
        n_users = chk.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        n_eu = chk.execute("SELECT COUNT(*) FROM email_updates").fetchone()[0]
        n_ea = chk.execute("SELECT COUNT(*) FROM email_attachments").fetchone()[0]
        n_se = chk.execute("SELECT COUNT(*) FROM shipment_events").fetchone()[0]
    finally:
        chk.close()
    log.info("  backup OK — size=%d bytes  (users=%d  email_updates=%d  email_attachments=%d  shipment_events=%d)",
             out.stat().st_size, n_users, n_eu, n_ea, n_se)
    return out


# ----- step 3: audit export -----

def step_audit_export(log: logging.Logger, ts: str) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORTS_DIR / f"post_sync_cleanup_audit_{ts}.xlsx"
    log.info("STEP 3: writing audit xlsx -> %s", out)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    con = sqlite3.connect(str(DB_PATH))
    try:
        cur = con.cursor()
        # Summary sheet
        sw = wb.create_sheet("_summary")
        sw.append(["table", "row_count", "category"])
        for t in SCOPED_TABLES_DELETE_ORDER:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError:
                n = "(missing)"
            sw.append([t, n, "scoped (will be cleared)"])
        for t in REQUIRE_EMPTY_BEFORE + PRESERVED_TABLES:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError:
                n = "(missing)"
            sw.append([t, n, "preserved / required-empty"])
        sw.append([])
        sw.append(["timestamp_utc", datetime.utcnow().isoformat()])

        for t in SCOPED_TABLES_DELETE_ORDER:
            try:
                cur.execute(f"SELECT * FROM {t}")
            except sqlite3.OperationalError:
                continue
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
            ws = wb.create_sheet(t[:31])
            ws.append(cols)
            for r in rows:
                ws.append([_cell(v) for v in r])
            log.info("  sheet '%s' — %d rows", t, len(rows))
        wb.save(out)
        log.info("  audit xlsx saved (%d bytes)", out.stat().st_size)
    finally:
        con.close()
    return out


# ----- step 4: archive new files -----

def step_archive_files(log: logging.Logger, ts: str) -> tuple[Path, int]:
    ARCHIVE_ROOT.mkdir(parents=True, exist_ok=True)
    target = ARCHIVE_ROOT / ts
    target.mkdir(parents=True, exist_ok=True)
    log.info("STEP 4: archiving %s -> %s", UPLOADS_DOCS, target)

    if not UPLOADS_DOCS.exists():
        log.info("  uploads/documents/ does not exist — nothing to move")
        return target, 0

    n = 0
    for child in list(UPLOADS_DOCS.iterdir()):
        dest = target / child.name
        try:
            shutil.move(str(child), str(dest))
            n += 1
        except Exception as e:
            log.error("  failed to move %s: %s", child, e)
            raise
    log.info("  moved %d items into %s", n, target)
    UPLOADS_DOCS.mkdir(parents=True, exist_ok=True)
    return target, n


# ----- step 5: scoped delete -----

def step_delete(log: logging.Logger) -> dict:
    log.info("STEP 5: deleting scoped tables (in dependency order)")
    con = sqlite3.connect(str(DB_PATH))
    con.isolation_level = None
    cur = con.cursor()
    cur.execute("BEGIN IMMEDIATE")
    try:
        deleted: dict[str, int] = {}
        for t in SCOPED_TABLES_DELETE_ORDER:
            try:
                before = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError as e:
                log.warning("  table %s missing — skipping (%s)", t, e)
                continue
            cur.execute(f"DELETE FROM {t}")
            after = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            log.info("  %-40s deleted %d rows (now %d)", t, before, after)
            if after != 0:
                raise RuntimeError(f"table {t} still has {after} rows after DELETE")
            deleted[t] = before

        try:
            for t in SCOPED_TABLES_DELETE_ORDER:
                cur.execute("DELETE FROM sqlite_sequence WHERE name=?", (t,))
            log.info("  sqlite_sequence reset for emptied tables")
        except sqlite3.OperationalError:
            log.info("  sqlite_sequence not present — skipped")

        cur.execute("COMMIT")
        log.info("  COMMIT OK — totals deleted: %s", deleted)
        return deleted
    except Exception:
        cur.execute("ROLLBACK")
        log.exception("  ROLLBACK due to error")
        raise
    finally:
        con.close()


# ----- step 6: verify -----

def step_verify(log: logging.Logger) -> dict:
    log.info("STEP 6: verifying final state")
    con = sqlite3.connect(str(DB_PATH))
    try:
        cur = con.cursor()
        post = {}
        for t in SCOPED_TABLES_DELETE_ORDER + REQUIRE_EMPTY_BEFORE:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError:
                continue
            post[t] = n
            log.info("  %-40s = %d %s", t, n, "OK" if n == 0 else "FAIL")
            if n != 0:
                raise RuntimeError(f"table {t} not empty (n={n})")
        for t in PRESERVED_TABLES:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                post[t] = n
                log.info("  %-40s = %d (preserved)", t, n)
            except sqlite3.OperationalError:
                log.warning("  preserved table %s missing", t)

        n_admin = cur.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]
        log.info("  admin users remaining = %d", n_admin)
        post["admin_users"] = n_admin
        if n_admin == 0:
            raise RuntimeError("no admin user remains — refusing to leave system unbootable")
        return post
    finally:
        con.close()


# ----- step 7: health check -----

def step_health(log: logging.Logger) -> int:
    log.info("STEP 7: running app health check")
    if not HEALTH_SCRIPT.exists():
        log.warning("  health script not found — skipping")
        return -1
    proc = subprocess.run(
        [sys.executable, str(HEALTH_SCRIPT)],
        cwd=str(BACKEND_ROOT),
        capture_output=True,
        text=True,
        timeout=120,
    )
    log.info("  exit code: %d", proc.returncode)
    for line in (proc.stdout or "").splitlines():
        log.info("  | %s", line)
    if proc.stderr:
        for line in proc.stderr.splitlines():
            log.warning("  ! %s", line)
    return proc.returncode


# ----- main -----

def main():
    args = parse_args()
    if args.confirm != CONFIRMATION_TOKEN:
        sys.stderr.write(
            "REFUSED: this is a destructive operation.\n"
            f"  Re-run with: python {sys.argv[0]} --confirm {CONFIRMATION_TOKEN}\n"
        )
        sys.exit(2)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOGS_DIR / f"post_sync_cleanup_{ts}.log"
    log = setup_logging(log_path)

    log.info("=" * 70)
    log.info("POST-SYNC CLEANUP")
    log.info("  timestamp = %s", ts)
    log.info("  db_path   = %s", DB_PATH)
    log.info("  log_path  = %s", log_path)
    log.info("=" * 70)

    try:
        step_preflight(log)
        backup_path = step_backup_db(log, ts)
        audit_path  = step_audit_export(log, ts)
        archive_path, n_files_moved = step_archive_files(log, ts)
        deleted = step_delete(log)
        post = step_verify(log)
        rc = step_health(log)
    except Exception as e:
        log.exception("ABORT — %s", e)
        sys.exit(3)

    log.info("=" * 70)
    log.info("CLEANUP COMPLETE")
    log.info("  backup db    : %s", backup_path)
    log.info("  audit xlsx   : %s", audit_path)
    log.info("  files moved  : %s  (%d items)", archive_path, n_files_moved)
    log.info("  log file     : %s", log_path)
    log.info("  health rc    : %s", rc)
    log.info("  deleted_rows : %s", deleted)
    log.info("  post_state   : %s", post)
    log.info("=" * 70)


if __name__ == "__main__":
    main()
