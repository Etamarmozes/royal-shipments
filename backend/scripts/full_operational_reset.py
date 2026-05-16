#!/usr/bin/env python
"""Full operational reset — backup, audit, archive, delete.

Usage:
    python backend/scripts/full_operational_reset.py --confirm FULL_RESET

Refuses to run without the literal "FULL_RESET" confirmation token.

What gets reset (operational data):
    - alerts
    - document_assignment_actions
    - document_assignment_qc_results
    - pending_document_updates
    - shipment_events
    - email_attachments      (and the underlying files are MOVED, NOT deleted)
    - extra_work_tasks
    - pending_containers
    - pending_shipments
    - email_updates
    - containers
    - shipments
    - import_batches

What is preserved (NOT touched):
    - users + roles + password hashes + last_login_at
    - document_assignment_rules (config / supplier rules)
    - All code and configuration files
    - Auth secrets, JWT settings
    - Gmail OAuth token (data/gmail_token.json)
    - Email-sync state cursors under data/state/

Outputs (all timestamped with %Y%m%d_%H%M%S):
    - DB backup:           backend/data/backups/royal_linen_before_full_reset_<ts>.db
    - Audit xlsx:          backend/exports/full_reset_audit_<ts>.xlsx
    - Log file:            backend/logs/full_reset_<ts>.log
    - File archive folder: backend/uploads/archive_before_reset/<ts>/

If anything fails the script aborts before deleting any rows.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime
from pathlib import Path

CONFIRMATION_TOKEN = "FULL_RESET"

# Resolve paths from this script's location, so it works regardless of cwd.
BACKEND_ROOT   = Path(__file__).resolve().parent.parent      # backend/
DB_PATH        = BACKEND_ROOT / "data" / "royal_linen.db"
BACKUPS_DIR    = BACKEND_ROOT / "data" / "backups"
EXPORTS_DIR    = BACKEND_ROOT / "exports"
LOGS_DIR       = BACKEND_ROOT / "logs"
UPLOADS_DOCS   = BACKEND_ROOT / "uploads" / "documents"
ARCHIVE_ROOT   = BACKEND_ROOT / "uploads" / "archive_before_reset"
HEALTH_SCRIPT  = BACKEND_ROOT / "scripts" / "check_app_health.py"

# Children-before-parents.  All operational tables are listed; tables NOT
# listed here (e.g. users, document_assignment_rules) are preserved.
OPERATIONAL_TABLES_DELETE_ORDER = [
    "alerts",
    "document_assignment_actions",
    "document_assignment_qc_results",
    "pending_document_updates",
    "shipment_events",
    "email_attachments",
    "extra_work_tasks",
    "pending_containers",
    "pending_shipments",
    "email_updates",
    "containers",
    "shipments",
    "import_batches",
]

PRESERVED_TABLES = [
    "users",
    "document_assignment_rules",
]


# ----- helpers -----

def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument(
        "--confirm",
        default="",
        help=f"Must equal '{CONFIRMATION_TOKEN}' to proceed.",
    )
    return ap.parse_args()


def setup_logging(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("full_reset")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    sh = logging.StreamHandler()
    sh.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


def _cell(v):
    """Coerce DB value into something openpyxl can write."""
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, bytes):
        try:
            return v.decode("utf-8", errors="replace")
        except Exception:
            return repr(v)
    try:
        return json.dumps(v, ensure_ascii=False)
    except Exception:
        return str(v)


# ----- step 1: backup DB -----

def step_backup_db(log: logging.Logger, ts: str) -> Path:
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    out = BACKUPS_DIR / f"royal_linen_before_full_reset_{ts}.db"
    log.info("STEP 1: backing up DB -> %s", out)

    if not DB_PATH.exists():
        raise RuntimeError(f"DB file does not exist: {DB_PATH}")

    src = sqlite3.connect(str(DB_PATH))
    try:
        dst = sqlite3.connect(str(out))
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()

    if not out.exists() or out.stat().st_size == 0:
        raise RuntimeError(f"backup file missing or empty: {out}")

    # Sanity: open backup and read users + shipments counts.
    chk = sqlite3.connect(str(out))
    try:
        n_users = chk.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        n_ships = chk.execute("SELECT COUNT(*) FROM shipments").fetchone()[0]
        n_docs  = chk.execute("SELECT COUNT(*) FROM email_attachments").fetchone()[0]
    finally:
        chk.close()
    log.info("  backup OK — size=%d bytes  (users=%d  shipments=%d  docs=%d)",
             out.stat().st_size, n_users, n_ships, n_docs)
    return out


# ----- step 2: audit export -----

def step_audit_export(log: logging.Logger, ts: str) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORTS_DIR / f"full_reset_audit_{ts}.xlsx"
    log.info("STEP 2: writing audit xlsx -> %s", out)

    from openpyxl import Workbook   # in requirements.txt

    wb = Workbook()
    wb.remove(wb.active)

    con = sqlite3.connect(str(DB_PATH))
    try:
        cur = con.cursor()

        # Summary sheet first
        sw = wb.create_sheet("_summary")
        sw.append(["table", "row_count", "category"])
        for t in OPERATIONAL_TABLES_DELETE_ORDER:
            try:
                cnt = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError:
                cnt = "(missing)"
            sw.append([t, cnt, "operational (will be cleared)"])
        for t in PRESERVED_TABLES:
            try:
                cnt = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError:
                cnt = "(missing)"
            sw.append([t, cnt, "preserved"])
        sw.append([])
        sw.append(["timestamp_utc", datetime.utcnow().isoformat()])
        sw.append(["db_path", str(DB_PATH)])

        # Per-table sheets — only operational data; users + rules excluded for privacy.
        for t in OPERATIONAL_TABLES_DELETE_ORDER:
            try:
                cur.execute(f"SELECT * FROM {t}")
            except sqlite3.OperationalError as e:
                log.warning("  table %s missing — sheet skipped (%s)", t, e)
                continue
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
            ws = wb.create_sheet(t[:31])    # Excel sheet name max 31 chars
            ws.append(cols)
            for r in rows:
                ws.append([_cell(v) for v in r])
            log.info("  sheet '%s' — %d rows", t, len(rows))
        wb.save(out)
        log.info("  audit xlsx saved (%d bytes)", out.stat().st_size)
    finally:
        con.close()
    return out


# ----- step 3: archive physical files -----

def step_archive_files(log: logging.Logger, ts: str) -> Path:
    ARCHIVE_ROOT.mkdir(parents=True, exist_ok=True)
    target = ARCHIVE_ROOT / ts
    target.mkdir(parents=True, exist_ok=True)
    log.info("STEP 3: archiving %s -> %s", UPLOADS_DOCS, target)

    if not UPLOADS_DOCS.exists():
        log.info("  uploads/documents/ does not exist — nothing to move")
        return target

    n_moved = 0
    for child in list(UPLOADS_DOCS.iterdir()):
        dest = target / child.name
        try:
            shutil.move(str(child), str(dest))
            n_moved += 1
        except Exception as e:
            log.error("  failed to move %s: %s", child, e)
            raise
    log.info("  moved %d items into %s", n_moved, target)

    # Recreate empty documents/ so the backend can write to it again.
    UPLOADS_DOCS.mkdir(parents=True, exist_ok=True)
    return target


# ----- step 4: delete operational rows -----

def step_delete_operational(log: logging.Logger) -> dict:
    log.info("STEP 4: deleting operational tables (in dependency order)")
    con = sqlite3.connect(str(DB_PATH))
    con.isolation_level = None
    cur = con.cursor()
    cur.execute("BEGIN IMMEDIATE")
    try:
        deleted: dict[str, int] = {}
        for t in OPERATIONAL_TABLES_DELETE_ORDER:
            try:
                before = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError as e:
                log.warning("  table %s missing — skipping (%s)", t, e)
                continue
            cur.execute(f"DELETE FROM {t}")
            after = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            log.info("  %-40s deleted %d rows (now %d)", t, before, after)
            if after != 0:
                raise RuntimeError(f"table {t} still has {after} rows after DELETE")
            deleted[t] = before

        # Reset autoincrement sequences ONLY for tables we just emptied.
        # Safe because there is nothing left to collide with.
        try:
            for t in OPERATIONAL_TABLES_DELETE_ORDER:
                cur.execute("DELETE FROM sqlite_sequence WHERE name=?", (t,))
            log.info("  sqlite_sequence reset for emptied tables")
        except sqlite3.OperationalError:
            log.info("  sqlite_sequence not present — skipped (no autoincrement)")

        cur.execute("COMMIT")
        log.info("  COMMIT OK — totals deleted: %s", deleted)
        return deleted
    except Exception:
        cur.execute("ROLLBACK")
        log.exception("  ROLLBACK due to error")
        raise
    finally:
        con.close()


# ----- step 5: verify post-state -----

def step_verify(log: logging.Logger) -> dict:
    log.info("STEP 5: verifying final state")
    con = sqlite3.connect(str(DB_PATH))
    try:
        cur = con.cursor()
        post = {}
        for t in OPERATIONAL_TABLES_DELETE_ORDER:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except sqlite3.OperationalError:
                continue
            post[t] = n
            log.info("  %-40s = %d %s", t, n, "OK" if n == 0 else "FAIL")
            if n != 0:
                raise RuntimeError(f"table {t} not empty (n={n})")

        for t in PRESERVED_TABLES:
            try:
                n = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                post[t] = n
                log.info("  %-40s = %d (preserved)", t, n)
            except sqlite3.OperationalError:
                log.warning("  preserved table %s missing", t)

        # Defensive: at least one admin user must remain or login is impossible.
        n_admin = cur.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]
        log.info("  admin users remaining = %d", n_admin)
        post["admin_users"] = n_admin
        if n_admin == 0:
            raise RuntimeError("no admin user remains — refusing to leave system unbootable")
        return post
    finally:
        con.close()


# ----- step 6: health check -----

def step_health_check(log: logging.Logger) -> int:
    log.info("STEP 6: running app health check")
    if not HEALTH_SCRIPT.exists():
        log.warning("  health script not found at %s — skipping", HEALTH_SCRIPT)
        return -1
    proc = subprocess.run(
        [sys.executable, str(HEALTH_SCRIPT)],
        cwd=str(BACKEND_ROOT),
        capture_output=True,
        text=True,
        timeout=120,
    )
    log.info("  exit code: %d", proc.returncode)
    for line in (proc.stdout or "").splitlines():
        log.info("  | %s", line)
    if proc.stderr:
        for line in proc.stderr.splitlines():
            log.warning("  ! %s", line)
    return proc.returncode


# ----- main -----

def main():
    args = parse_args()
    if args.confirm != CONFIRMATION_TOKEN:
        sys.stderr.write(
            "REFUSED: this is a destructive operational reset.\n"
            f"  Re-run with: python {sys.argv[0]} --confirm {CONFIRMATION_TOKEN}\n"
        )
        sys.exit(2)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOGS_DIR / f"full_reset_{ts}.log"
    log = setup_logging(log_path)

    log.info("=" * 70)
    log.info("FULL OPERATIONAL RESET")
    log.info("  timestamp = %s", ts)
    log.info("  db_path   = %s", DB_PATH)
    log.info("  log_path  = %s", log_path)
    log.info("=" * 70)

    try:
        backup_path  = step_backup_db(log, ts)
        audit_path   = step_audit_export(log, ts)
        archive_path = step_archive_files(log, ts)
        deleted      = step_delete_operational(log)
        post         = step_verify(log)
        rc           = step_health_check(log)
    except Exception as e:
        log.exception("ABORT — %s", e)
        log.error("If the script aborted before STEP 4, no DB rows were deleted.")
        log.error("If it aborted during STEP 4, the transaction was rolled back.")
        log.error("Backup file location (if backup completed): %s", BACKUPS_DIR)
        sys.exit(3)

    log.info("=" * 70)
    log.info("RESET COMPLETE")
    log.info("  backup db    : %s", backup_path)
    log.info("  audit xlsx   : %s", audit_path)
    log.info("  files moved  : %s", archive_path)
    log.info("  log file     : %s", log_path)
    log.info("  health rc    : %s", rc)
    log.info("  deleted_rows : %s", deleted)
    log.info("  post_state   : %s", post)
    log.info("=" * 70)


if __name__ == "__main__":
    main()
