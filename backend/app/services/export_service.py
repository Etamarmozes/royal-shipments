from datetime import datetime
from io import BytesIO
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..models import (
    Shipment, Container, ExtraWorkTask, EmailUpdate,
    PendingShipment, Alert, ShipmentEvent,
)
from . import dashboard_service, forecast_service


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _set_headers(ws, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right")
    ws.sheet_view.rightToLeft = True


def _write_dashboard(ws, db: Session):
    kpis = dashboard_service.kpis(db)
    ws.append(["מדד", "ערך"])
    _set_headers(ws, ["מדד", "ערך"])
    for k, v in kpis.items():
        ws.append([k, str(v) if v is not None else ""])


def _write_shipments(ws, db: Session, archived: bool = False):
    headers = [
        "SHP", "ספק", "תיאור", "מקור", "שלב", "ETD",
        "ETA לארץ", "ETA נמל", "ETA מחסן",
        "בוקינג", "BOL", "Invoice", "PO",
        "עיכוב", "סיבת עיכוב", "ניירת מלאה", "תוספת עבודה",
        "מקור עדכון אחרון", "עדכון אחרון",
    ]
    _set_headers(ws, headers)
    rows = (
        db.query(Shipment).filter(Shipment.archived == archived)
        .order_by(Shipment.id.asc()).all()
    )
    for s in rows:
        ws.append([
            s.shp_id, s.supplier, s.goods_description, s.origin_country,
            s.current_stage, s.etd, s.eta_israel, s.eta_port, s.eta_warehouse,
            s.booking_number, s.bol_number, s.invoice_number, s.po_number,
            "כן" if s.delay_status else "לא", s.delay_reason or "",
            "כן" if s.paperwork_complete else "לא",
            "כן" if s.extra_work_required else "לא",
            s.last_update_source or "", s.updated_at,
        ])


def _write_containers(ws, db: Session):
    headers = [
        "מכולה", "סוג", "SHP", "ספק", "תיאור", "CBM", "קופסאות", "משקל",
        "סטטוס", "ETA לארץ", "ETA נמל", "ETA מחסן",
        "עדיפות", "תוספת עבודה",
    ]
    _set_headers(ws, headers)
    rows = db.query(Container).options(joinedload(Container.shipment)).all()
    for c in rows:
        s = c.shipment
        ws.append([
            c.container_number, c.container_type,
            s.shp_id if s else "", s.supplier if s else "",
            s.goods_description if s else "",
            c.cbm, c.boxes_total, c.gross_weight_kg,
            c.container_status,
            c.eta_israel or (s.eta_israel if s else None),
            c.eta_port or (s.eta_port if s else None),
            c.eta_warehouse or (s.eta_warehouse if s else None),
            c.unloading_priority,
            "כן" if c.extra_work_required else "לא",
        ])


def _write_forecast(ws, db: Session):
    headers = ["שבוע", "מתאריך", "עד תאריך", "מכולות לארץ",
               "מכולות לנמל", "מכולות למחסן", "CBM", "משקל",
               "קופסאות", "ספקים", "סטטוס עומס"]
    _set_headers(ws, headers)
    weeks = forecast_service.forecast_8_weeks(db)
    for w in weeks:
        ws.append([
            w["week_label"], w["week_start"], w["week_end"],
            w["containers_arriving_israel"], w["containers_arriving_port"],
            w["containers_arriving_warehouse"], w["cbm_total"],
            w["weight_total_kg"], w["boxes_total"],
            ", ".join(w["suppliers"]), w["load_status"],
        ])


def _write_email_updates(ws, db: Session):
    headers = ["#", "תאריך", "שולח", "נושא", "סוג זיהוי", "משלוח משויך",
               "ביטחון", "סטטוס"]
    _set_headers(ws, headers)
    rows = db.query(EmailUpdate).order_by(EmailUpdate.id.desc()).all()
    for u in rows:
        s = None
        if u.detected_shipment_id:
            s = db.query(Shipment).filter(Shipment.id == u.detected_shipment_id).first()
        ws.append([
            u.id, u.received_at, u.sender, u.subject,
            u.detection_type, s.shp_id if s else "",
            u.confidence_score, u.status,
        ])


def _write_pending_shipments(ws, db: Session):
    headers = ["#", "ספק", "תיאור", "ETA לארץ", "בוקינג", "BOL",
               "Invoice", "PO", "ביטחון", "סטטוס", "נוצר"]
    _set_headers(ws, headers)
    for ps in db.query(PendingShipment).order_by(PendingShipment.id.desc()).all():
        ws.append([
            ps.id, ps.detected_supplier, ps.detected_goods_description,
            ps.detected_eta_israel, ps.detected_booking_number,
            ps.detected_bol_number, ps.detected_invoice_number,
            ps.detected_po_number, ps.confidence_score,
            ps.status, ps.created_at,
        ])


def _write_extra_work(ws, db: Session):
    headers = ["#", "SHP", "סוג עבודה", "אחראי", "סטטוס",
               "התחלה צפויה", "סיום צפוי", "סיום בפועל",
               "מוכן להפצה (משוער)", "עיכוב", "סיבת עיכוב"]
    _set_headers(ws, headers)
    for t in db.query(ExtraWorkTask).order_by(ExtraWorkTask.id.desc()).all():
        s = db.query(Shipment).filter(Shipment.id == t.shipment_id).first()
        ws.append([
            t.id, s.shp_id if s else "", t.work_type, t.responsible_party,
            t.work_status, t.expected_start_date, t.expected_end_date,
            t.actual_end_date, t.ready_for_distribution_estimated_date,
            "כן" if t.delay_status else "לא", t.delay_reason or "",
        ])


def _write_alerts(ws, db: Session):
    headers = ["#", "סוג", "חומרה", "כותרת", "תיאור", "SHP", "טופל", "תאריך"]
    _set_headers(ws, headers)
    for a in db.query(Alert).order_by(Alert.id.desc()).all():
        s = None
        if a.shipment_id:
            s = db.query(Shipment).filter(Shipment.id == a.shipment_id).first()
        ws.append([
            a.id, a.alert_type, a.severity, a.title, a.description or "",
            s.shp_id if s else "",
            "כן" if a.resolved else "לא", a.created_at,
        ])


def _write_history(ws, db: Session):
    headers = ["SHP", "ספק", "תיאור", "מדינה", "ETA לארץ",
               "הגעה בפועל", "ימי איחור", "תוספת עבודה", "הסתיים בתאריך"]
    _set_headers(ws, headers)
    rows = db.query(Shipment).filter(Shipment.archived == True).all()  # noqa: E712
    for s in rows:
        delay_days = None
        if s.eta_israel and s.actual_arrival_israel:
            delay_days = (s.actual_arrival_israel - s.eta_israel).days
        ws.append([
            s.shp_id, s.supplier, s.goods_description, s.origin_country,
            s.eta_israel, s.actual_arrival_israel, delay_days,
            "כן" if s.extra_work_required else "לא", s.completed_at,
        ])


def _write_change_log(ws, db: Session):
    headers = ["#", "תאריך", "ישות", "מזהה", "פעולה", "שדה",
               "ערך ישן", "ערך חדש", "מי", "מקור"]
    _set_headers(ws, headers)
    for e in db.query(ShipmentEvent).order_by(ShipmentEvent.id.desc()).limit(2000).all():
        ws.append([
            e.id, e.changed_at, e.entity_type, e.entity_id,
            e.action_type, e.field_changed or "",
            e.old_value or "", e.new_value or "",
            e.changed_by or "", e.source,
        ])


def build_workbook(db: Session) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Dashboard")
    _write_dashboard(ws, db)

    ws = wb.create_sheet("Active Shipments")
    _write_shipments(ws, db, archived=False)

    ws = wb.create_sheet("Containers")
    _write_containers(ws, db)

    ws = wb.create_sheet("8 Weeks Forecast")
    _write_forecast(ws, db)

    ws = wb.create_sheet("Email Updates")
    _write_email_updates(ws, db)

    ws = wb.create_sheet("Pending Shipments")
    _write_pending_shipments(ws, db)

    ws = wb.create_sheet("Extra Work")
    _write_extra_work(ws, db)

    ws = wb.create_sheet("Alerts")
    _write_alerts(ws, db)

    ws = wb.create_sheet("History")
    _write_history(ws, db)

    ws = wb.create_sheet("Change Log")
    _write_change_log(ws, db)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
