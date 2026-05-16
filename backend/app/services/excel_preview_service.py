"""Excel file → JSON sheets for inline preview in the UI.

Supports:
- .xlsx / .xlsm (openpyxl)
- .xls (xlrd if available, otherwise return informative error)

Returns a structured dict the React component can render as tabs+tables.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

log = logging.getLogger("excel_preview")

MAX_ROWS_PER_SHEET = 200          # hard cap on what we send to the UI
MAX_COLS_PER_SHEET = 30
MAX_CELL_CHARS = 500


def _trim_cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    if len(s) > MAX_CELL_CHARS:
        return s[:MAX_CELL_CHARS] + "…"
    return s


def _preview_xlsx(path: Path) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed"}
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        log.warning("xlsx open failed: %s", e)
        return {"error": f"לא הצלחנו לפתוח את הקובץ: {e}"}

    sheets: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        rows: List[List[Any]] = []
        actual_row_count = 0
        actual_col_count = 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            actual_row_count = r_idx + 1
            actual_col_count = max(actual_col_count, len(row))
            if r_idx >= MAX_ROWS_PER_SHEET:
                continue  # keep counting but don't store
            row_trimmed = [_trim_cell(c) for c in row[:MAX_COLS_PER_SHEET]]
            rows.append(row_trimmed)
        sheets.append({
            "name": ws.title,
            "rows": rows,
            "row_count": actual_row_count,
            "col_count": actual_col_count,
            "truncated": actual_row_count > MAX_ROWS_PER_SHEET or actual_col_count > MAX_COLS_PER_SHEET,
        })
    wb.close()
    return {"format": "xlsx", "sheets": sheets}


def _preview_xls(path: Path) -> Dict[str, Any]:
    try:
        import xlrd  # not in requirements; may fail
    except ImportError:
        return {
            "error": (
                "תצוגה מקדימה של xls (פורמט ישן) אינה זמינה. "
                "ניתן להוריד את הקובץ ולפתוח באקסל."
            ),
            "format": "xls",
            "sheets": [],
        }
    try:
        wb = xlrd.open_workbook(str(path))
    except Exception as e:
        return {"error": f"לא הצלחנו לפתוח: {e}"}
    sheets = []
    for ws in wb.sheets():
        rows: List[List[Any]] = []
        for r_idx in range(min(ws.nrows, MAX_ROWS_PER_SHEET)):
            row = ws.row_values(r_idx, end_colx=MAX_COLS_PER_SHEET)
            rows.append([_trim_cell(c) for c in row])
        sheets.append({
            "name": ws.name,
            "rows": rows,
            "row_count": ws.nrows,
            "col_count": ws.ncols,
            "truncated": ws.nrows > MAX_ROWS_PER_SHEET or ws.ncols > MAX_COLS_PER_SHEET,
        })
    return {"format": "xls", "sheets": sheets}


def preview(path: Path) -> Dict[str, Any]:
    """Detect format from the on-disk magic bytes (more reliable than extension)."""
    if not path.exists():
        return {"error": "Missing on disk"}
    if path.stat().st_size == 0:
        return {"error": "Empty file"}
    with open(path, "rb") as f:
        head = f.read(8)
    if head[:4] == b"PK\x03\x04":
        return _preview_xlsx(path)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return _preview_xls(path)
    # Fallback by extension
    suf = path.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return _preview_xlsx(path)
    if suf == ".xls":
        return _preview_xls(path)
    return {"error": f"לא קובץ Excel ידוע (signature: {head.hex()})"}
