"""Detect which kind of shipment Excel file the user uploaded.

Three formats supported (preview-only, no DB writes):

  royal_linen_template  — our own template (cols A..AN, header row 2,
                          technical-key row 3). Existing parser handles it.
  icl                   — "ICL / Israel Cargo Logistics" report.
                          Sheet "Open Import Orders & Files".
                          Title rows 1..14, header row 15, data 16..22.
  eli_line              — "Eli Line" container list.
                          Sheet "גיליון1". Header row 1, ASHDOD section row 2,
                          data rows 3..18.

Detection runs on the loaded workbook — no DB calls. Returns the format
key + the active sheet name + the row index of the actual header.
"""
from __future__ import annotations

from io import BytesIO
from typing import Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _str(v) -> str:
    return ("" if v is None else str(v)).strip()


def _row_text(ws, row: int, max_col: int = 20) -> str:
    """Concatenate the visible cell text for one row (lowercased)."""
    parts = []
    for c in range(1, max_col + 1):
        v = ws.cell(row, c).value
        if v is not None:
            parts.append(str(v).strip())
    return " ".join(parts).lower()


def detect_format(content: bytes) -> Tuple[str, dict]:
    """Returns (format_key, info). format_key is one of:
        - "royal_linen_template"
        - "icl"
        - "eli_line"
        - "unknown"
    info dict carries: sheet_name, header_row, source_provider, notes.
    """
    try:
        wb: Workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        return "unknown", {"error": f"שגיאה בקריאת הקובץ: {e}"}

    # ---- 1. ICL detection ----
    # Sheet name "Open Import Orders & Files" (sometimes "&" replaced) OR
    # row 15 contains "ICL File" / "Sho List" / "Customs File"
    for ws in wb.worksheets:
        title_low = ws.title.lower()
        if ("open import" in title_low or "open_import" in title_low
                or "import orders" in title_low):
            return "icl", {
                "sheet_name": ws.title,
                "header_row": 15,
                "data_first_row": 16,
                "source_provider": "ICL",
                "notes": "ICL standard format. Title rows 1-14 ignored.",
            }

    # Fallback: scan first 20 rows for an ICL header signature
    for ws in wb.worksheets:
        for r in range(1, min(21, ws.max_row + 1)):
            text = _row_text(ws, r, 16)
            if "icl file" in text or "sho list" in text or ("customs file" in text and "po number" in text):
                return "icl", {
                    "sheet_name": ws.title,
                    "header_row": r,
                    "data_first_row": r + 1,
                    "source_provider": "ICL",
                    "notes": f"ICL header detected at row {r}.",
                }

    # ---- 2. Eli Line detection ----
    # Hebrew column header signature: "תאריך הגעה משוער" + "JOB" + "MARKS"
    for ws in wb.worksheets:
        # Row 1 is the header in Eli Line
        text = _row_text(ws, 1, 12)
        if "תאריך הגעה משוער" in text and "job" in text:
            return "eli_line", {
                "sheet_name": ws.title,
                "header_row": 1,
                "data_first_row": 3,   # skip row 2 (ASHDOD section)
                "source_provider": "Eli Line",
                "notes": "Eli Line standard format. Row 2 (ASHDOD) is a section header — skipped.",
            }

    # ---- 3. Royal Linen template detection ----
    # Look for our technical-key row containing "supplier_name" + "shipment_reference"
    for ws in wb.worksheets:
        for r in range(1, min(11, ws.max_row + 1)):
            text = _row_text(ws, r, 40)
            if "supplier_name" in text and "shipment_reference" in text:
                return "royal_linen_template", {
                    "sheet_name": ws.title,
                    "header_row": r,
                    "data_first_row": r + 2,
                    "source_provider": "Royal Linen Template",
                    "notes": "Royal Linen template — technical key row.",
                }

    return "unknown", {
        "sheet_name": wb.worksheets[0].title if wb.worksheets else None,
        "notes": "Format not recognised. Upload one of: ICL, Eli Line, "
                 "or our official shipment_import_template.xlsx.",
    }
