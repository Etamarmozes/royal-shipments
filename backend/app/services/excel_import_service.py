"""Excel shipment import — template generator + parse + dedup + apply.

The flow is intentionally two-step (preview then apply) so the user can
review every row before it touches the database. Nothing here writes
shipment data without an explicit approved-rows list.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from ..config import BASE_DIR
from ..models import Container, Shipment
from . import event_service, shipment_service

log = logging.getLogger("excel-import")

EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)
TEMPLATE_PATH = EXPORT_DIR / "shipment_import_template.xlsx"


# =====================================================================
# Column schema — single source of truth for the template + parser.
#
# Each entry: (column_key, hebrew_label, required, type, sample, comment)
# type ∈ {"str", "int", "float", "date", "bool"}
# =====================================================================

COLUMNS: List[Tuple[str, str, bool, str, Any, str]] = [
    # ---- Identity ----
    ("shipment_reference",  "מס׳ משלוח (SHP-XXX)",  False, "str",  "SHP-100",
     "אופציונלי. אם ריק, המערכת תקצה SHP חדש אוטומטית."),
    ("supplier_name",       "ספק",                  True,  "str",  "Nandan Terry",
     "חובה."),
    ("brand",               "מותג",                 False, "str",  "Royal",            ""),
    ("category",            "קטגוריה",              False, "str",  "מגבות",
     "מתוך הרשימה המבוקרת."),
    ("purchase_order_number","PO",                  False, "str",  "PO-2026-001",      ""),
    ("invoice_number",      "Invoice",              False, "str",  "INV-77321",        ""),
    ("packing_list_number", "Packing List No.",     False, "str",  "PL-77321",         ""),
    ("bill_of_lading_number","BL / BOL",            False, "str",  "MAEU123456789",
     "מקש ייחודי לזיהוי משלוח חוזר."),
    ("forwarder_name",      "עמיל / מעביר",         False, "str",  "DHL",              ""),
    ("shipping_company",    "חברת שילוח",           False, "str",  "Maersk",           ""),
    ("vessel_name",         "אוניה",                False, "str",  "Maersk Salina",    ""),
    ("container_number",    "מספר מכולה",           False, "str",  "MAEU1234567",
     "מקש ייחודי. מכולות מרובות במשלוח אחד — שורה לכל מכולה עם אותו shipment_reference."),
    ("container_type",      "סוג מכולה",            False, "str",  "40HC",             ""),

    # ---- Geography ----
    ("origin_country",      "מדינת מקור",           False, "str",  "India",            ""),
    ("origin_port",         "נמל יציאה",            False, "str",  "Mundra",           ""),
    ("destination_port",    "נמל יעד",              False, "str",  "Ashdod",           ""),
    ("destination_warehouse","מחסן יעד",            False, "str",  "Royal Linen MC",   ""),
    ("incoterm",            "Incoterm",             False, "str",  "FOB",
     "FOB / CIF / EXW / DAP / DDP …"),
    ("shipment_status",     "סטטוס",                False, "str",  "in_transit",
     "ערכים מומלצים: ordered / in_transit / arrived / received / delayed / cancelled"),

    # ---- Dates (ISO YYYY-MM-DD) ----
    ("etd",                 "ETD",                  False, "date", "2026-06-01",
     "תאריך יציאה. פורמט YYYY-MM-DD."),
    ("eta_port",            "ETA נמל",              False, "date", "2026-06-22",       ""),
    ("eta_warehouse",       "ETA מחסן",             False, "date", "2026-06-26",       ""),
    ("actual_arrival_port", "הגעה בפועל לנמל",      False, "date", "",
     "השאר ריק עד שהמכולה מגיעה."),
    ("actual_arrival_warehouse","הגעה בפועל למחסן", False, "date","",                  ""),

    # ---- Commercial / operational ----
    ("number_of_cartons",   "כמות קרטונים",         False, "int", 940,                 ""),
    ("number_of_pallets",   "כמות משטחים (אם ידוע)",False, "int", 22,
     "אם ריק, המערכת תחשב מהמידות."),
    ("gross_weight",        "משקל ברוטו (ק״ג)",     False, "float",13500,              ""),
    ("cbm",                 "CBM",                  False, "float",65.4,               ""),
    ("shipment_value",      "ערך משלוח",            False, "float",54200,              ""),
    ("currency",            "מטבע",                 False, "str",  "USD",
     "USD / EUR / ILS / …"),
    ("payment_status",      "סטטוס תשלום",          False, "str",  "paid",
     "paid / partial / unpaid / pending"),
    ("customs_status",      "סטטוס מכס",            False, "str",  "released",
     "pending / released / held"),
    ("documents_status",    "סטטוס מסמכים",         False, "str",  "complete",
     "complete / partial / missing"),

    # ---- Document tracking (booleans) ----
    ("invoice_received",    "Invoice התקבל?",       False, "bool", "yes",
     "yes / no  או  true / false"),
    ("packing_list_received","Packing List התקבל?", False, "bool", "yes",              ""),
    ("bl_received",         "BL התקבל?",            False, "bool", "no",               ""),
    ("certificate_received","תעודה התקבלה?",        False, "bool", "no",               ""),
    ("other_documents_missing","מסמכים חסרים אחרים",False, "str",  "",                  ""),

    # ---- Free text ----
    ("notes",               "הערות",                False, "str",  "דחוף — מעבדה",      ""),
    ("internal_owner",      "אחראי פנים-ארגוני",    False, "str",  "Jacob",            ""),
    ("priority",            "עדיפות",               False, "str",  "high",
     "low / normal / high / urgent"),
]


# =====================================================================
# Template generation
# =====================================================================

def build_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    ws.sheet_view.rightToLeft = True

    # ---- Row 1: instructions ----
    ws.cell(1, 1, value=(
        "הוראות: מלא שורה לכל מכולה. "
        "אם משלוח כולל יותר ממכולה אחת — שתי שורות עם אותו shipment_reference. "
        "שדות חובה מסומנים בכתום. תאריכים: YYYY-MM-DD. "
        "השורות לדוגמה (3-5) — מחק לפני שמירה."
    ))
    ws.cell(1, 1).font = Font(bold=True, color="334155", size=11)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="F1F5F9")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 36
    ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="right")

    # ---- Row 2: header ----
    REQ_FILL = PatternFill("solid", fgColor="FED7AA")  # orange-200
    OPT_FILL = PatternFill("solid", fgColor="E2E8F0")  # slate-200
    HDR_FONT = Font(bold=True, color="0F172A", size=10)
    for idx, (key, label, required, kind, sample, comment) in enumerate(COLUMNS, start=1):
        c = ws.cell(2, idx, value=label)
        c.font = HDR_FONT
        c.fill = REQ_FILL if required else OPT_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if comment or kind != "str":
            note_parts = []
            if required: note_parts.append("שדה חובה.")
            note_parts.append(f"מפתח טכני: {key}")
            note_parts.append(f"סוג: {kind}")
            if comment: note_parts.append(comment)
            c.comment = Comment(text="\n".join(note_parts), author="Royal Linen")
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(label) + 4)
    ws.row_dimensions[2].height = 38

    # ---- Row 3: technical key (smaller, gray, for the parser) ----
    KEY_FONT = Font(italic=True, color="64748B", size=8)
    for idx, (key, *_rest) in enumerate(COLUMNS, start=1):
        c = ws.cell(3, idx, value=key)
        c.font = KEY_FONT
        c.alignment = Alignment(horizontal="center")

    # ---- Sample rows (5..7) ----
    samples = [
        # Sample 1: full multi-container shipment, container 1
        {
            "shipment_reference": "SAMPLE-1", "supplier_name": "Nandan Terry",
            "category": "מגבות", "purchase_order_number": "PO-2026-001",
            "invoice_number": "INV-77321", "packing_list_number": "PL-77321",
            "bill_of_lading_number": "MAEU123456789",
            "forwarder_name": "DHL", "shipping_company": "Maersk",
            "vessel_name": "Maersk Salina",
            "container_number": "MAEU1234567", "container_type": "40HC",
            "origin_country": "India", "origin_port": "Mundra",
            "destination_port": "Ashdod", "destination_warehouse": "Royal Linen MC",
            "incoterm": "FOB", "shipment_status": "in_transit",
            "etd": "2026-06-01", "eta_port": "2026-06-22", "eta_warehouse": "2026-06-26",
            "number_of_cartons": 940, "number_of_pallets": 22,
            "gross_weight": 13500, "cbm": 65.4,
            "shipment_value": 54200, "currency": "USD",
            "payment_status": "paid", "customs_status": "pending",
            "documents_status": "partial",
            "invoice_received": "yes", "packing_list_received": "yes",
            "bl_received": "no", "certificate_received": "no",
            "notes": "דוגמה — מחק לפני שמירה", "internal_owner": "Jacob",
            "priority": "high",
        },
        # Sample 2: SAME shipment, second container
        {
            "shipment_reference": "SAMPLE-1", "supplier_name": "Nandan Terry",
            "container_number": "MAEU7654321", "container_type": "40HC",
            "number_of_cartons": 800, "cbm": 60.1,
            "notes": "שורה 2 של אותו משלוח — מאותו ספק / אותה הזמנה",
        },
        # Sample 3: standalone single-container shipment
        {
            "shipment_reference": "SAMPLE-2", "supplier_name": "Polder",
            "category": "כלי בית",
            "purchase_order_number": "PO-2026-002",
            "bill_of_lading_number": "ZIMUNYC1234567",
            "container_number": "ZIMU9988776", "container_type": "20GP",
            "origin_country": "China", "origin_port": "Shanghai",
            "destination_port": "Haifa", "destination_warehouse": "Royal Linen MC",
            "etd": "2026-07-10", "eta_port": "2026-08-08",
            "number_of_cartons": 320, "cbm": 22.0,
            "shipment_value": 18000, "currency": "USD",
            "shipment_status": "ordered",
            "notes": "shipment_reference ריק — המערכת תקצה SHP אוטומטית",
            "priority": "normal",
        },
    ]
    SAMPLE_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber-100
    for row_offset, row_data in enumerate(samples, start=5):
        for idx, (key, *_rest) in enumerate(COLUMNS, start=1):
            v = row_data.get(key, "")
            c = ws.cell(row_offset, idx, value=v)
            c.fill = SAMPLE_FILL

    # ---- Data validation: status, priority, currency, booleans ----
    n_max = 1000  # validate first 1000 rows
    def add_dv(formula1: str, col_idx: int, allow_blank: bool = True):
        dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
        first = get_column_letter(col_idx)
        dv.add(f"{first}5:{first}{n_max}")
        ws.add_data_validation(dv)

    col_idx = {key: i for i, (key, *_rest) in enumerate(COLUMNS, start=1)}
    add_dv('"ordered,in_transit,arrived,received,delayed,cancelled"',
           col_idx["shipment_status"])
    add_dv('"low,normal,high,urgent"', col_idx["priority"])
    add_dv('"USD,EUR,ILS,CNY,INR"', col_idx["currency"])
    add_dv('"yes,no,true,false,1,0"', col_idx["invoice_received"])
    add_dv('"yes,no,true,false,1,0"', col_idx["packing_list_received"])
    add_dv('"yes,no,true,false,1,0"', col_idx["bl_received"])
    add_dv('"yes,no,true,false,1,0"', col_idx["certificate_received"])
    add_dv('"paid,partial,unpaid,pending"', col_idx["payment_status"])
    add_dv('"pending,released,held"', col_idx["customs_status"])
    add_dv('"complete,partial,missing"', col_idx["documents_status"])

    # ---- Freeze top 3 rows, autofilter on the data range ----
    ws.freeze_panes = "A4"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A2:{last_col}{n_max}"

    # ---- README sheet ----
    rs = wb.create_sheet("README")
    rs.sheet_view.rightToLeft = True
    readme_lines = [
        ("Royal Linen — Shipment Import Template", True),
        ("", False),
        ("הוראות שימוש:", True),
        ("1. עיין בשורות 5-7 (כתום בהיר) — דוגמאות. מחק אותן לפני שמירה.", False),
        ("2. מלא שורה לכל מכולה.", False),
        ("3. אם משלוח כולל יותר ממכולה אחת — שורות מרובות עם אותו shipment_reference.", False),
        ("4. אם תשאיר shipment_reference ריק, המערכת תקצה SHP-XXX חדש.", False),
        ("5. שדות חובה מסומנים בכתום (supplier_name).", False),
        ("6. תאריכים בפורמט YYYY-MM-DD.", False),
        ("7. שמור את הקובץ והעלה אותו במסך 'ייבוא מאקסל'.", False),
        ("8. תקבל מסך תצוגה מקדימה — אשר לפני שזה נכנס למסד.", False),
        ("", False),
        ("Dedup keys (זיהוי כפילויות):", True),
        ("• shipment_reference (אם קיים)", False),
        ("• container_number", False),
        ("• bill_of_lading_number", False),
        ("• invoice_number", False),
        ("", False),
        ("פעולות בעת כפילות:", True),
        ("• create — צור חדש (אם הזיהוי לא נמצא)", False),
        ("• update — עדכן את הקיים (זיהוי תואם)", False),
        ("• skip — דלג על השורה", False),
    ]
    for r, (text, bold) in enumerate(readme_lines, start=1):
        c = rs.cell(r, 1, value=text)
        c.font = Font(bold=bold, size=12 if bold else 10)
        c.alignment = Alignment(horizontal="right", wrap_text=True)
    rs.column_dimensions["A"].width = 100

    return wb


def write_template_to_disk() -> Path:
    wb = build_template_workbook()
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


# =====================================================================
# Parsing + validation
# =====================================================================

def _to_str(v) -> Optional[str]:
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _to_int(v) -> Optional[int]:
    if v in (None, ""): return None
    try:
        return int(float(v))
    except Exception:
        return None


def _to_float(v) -> Optional[float]:
    if v in (None, ""): return None
    try:
        return float(v)
    except Exception:
        return None


def _to_date(v) -> Optional[date]:
    if v in (None, ""): return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


def _to_bool(v) -> Optional[bool]:
    if v in (None, ""): return None
    if isinstance(v, bool): return v
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1", "כן"):  return True
    if s in ("no", "n", "false", "0", "לא"):  return False
    return None


_COERCERS = {"str": _to_str, "int": _to_int, "float": _to_float,
             "date": _to_date, "bool": _to_bool}


def parse_workbook(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Return (rows, errors). Header row = 2, key row = 3, data starts row 5
    in the official template; we accept anywhere from row 2+ for flexibility:
    we look for the technical-key row (matches our COLUMNS keys) and start
    parsing one row after it.
    """
    errors: List[str] = []
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        return [], [f"שגיאה בקריאת הקובץ: {e}"]

    ws = wb.active

    # Locate the technical-key row by scanning the first 10 rows
    expected_keys = {key for key, *_ in COLUMNS}
    key_row_idx = None
    key_to_col: Dict[str, int] = {}
    for r in range(1, 11):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        normalized = [(str(v).strip() if v else "") for v in row_vals]
        matched = sum(1 for v in normalized if v in expected_keys)
        if matched >= 5:  # 5+ keys matched → that's the key row
            key_row_idx = r
            for col, val in enumerate(normalized, start=1):
                if val in expected_keys:
                    key_to_col[val] = col
            break
    if key_row_idx is None:
        # Fallback: try to use the Hebrew label row (row 2 by convention)
        label_to_key = {label: key for key, label, *_ in COLUMNS}
        for r in range(1, 11):
            row_vals = [str(ws.cell(r, c).value or "").strip()
                        for c in range(1, ws.max_column + 1)]
            matched = sum(1 for v in row_vals if v in label_to_key)
            if matched >= 5:
                key_row_idx = r
                for col, val in enumerate(row_vals, start=1):
                    k = label_to_key.get(val)
                    if k:
                        key_to_col[k] = col
                break

    if key_row_idx is None or not key_to_col:
        return [], [
            "לא זוהתה כותרת תקנית. השתמש בתבנית הרשמית "
            "(הורד מ-/import/excel/template) — שורה 3 חייבת להכיל את "
            "המפתחות הטכניים (supplier_name, container_number, …)."
        ]

    rows: List[Dict[str, Any]] = []
    coercers_by_key = {key: _COERCERS[kind] for key, _, _, kind, *_ in COLUMNS}
    required_keys = [key for key, _, req, *_ in COLUMNS if req]

    for r in range(key_row_idx + 1, ws.max_row + 1):
        raw = {key: ws.cell(r, col).value for key, col in key_to_col.items()}
        # Skip empty rows
        if all(v in (None, "") for v in raw.values()):
            continue
        # Skip sample rows. We mark every sample with "SAMPLE-" in
        # shipment_reference; we also check supplier_name as a backup
        # in case the user kept the shipment_reference but blanked it.
        sr = str(raw.get("shipment_reference") or "").strip().upper()
        sup = str(raw.get("supplier_name") or "").strip()
        if sr.startswith("SAMPLE"):
            continue
        # The supplier "Nandan Terry" / "Polder" alone are NOT a marker
        # since these may be real suppliers — only filter when sup is
        # clearly a placeholder.
        if sup.lower() in ("sample", "demo", "placeholder"):
            continue

        coerced: Dict[str, Any] = {}
        row_errors: List[str] = []
        for key, val in raw.items():
            try:
                coerced[key] = coercers_by_key[key](val)
            except Exception as e:
                row_errors.append(f"{key}: {e}")

        # Required-field check
        for rk in required_keys:
            if not coerced.get(rk):
                row_errors.append(f"שדה חובה ריק: {rk}")

        coerced["_row"] = r
        coerced["_errors"] = row_errors
        rows.append(coerced)

    return rows, errors


# =====================================================================
# Dedup
# =====================================================================

def find_duplicate_match(db: Session, row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Return the matched shipment summary (id, shp_id, source_field) if any
    of the dedup keys hits an existing row; else None."""
    sr = row.get("shipment_reference")
    if sr:
        s = db.query(Shipment).filter(Shipment.shp_id == sr).first()
        if s:
            return {"id": s.id, "shp_id": s.shp_id, "matched_by": "shipment_reference"}
    bl = row.get("bill_of_lading_number")
    if bl:
        s = db.query(Shipment).filter(Shipment.bol_number == bl).first()
        if s:
            return {"id": s.id, "shp_id": s.shp_id, "matched_by": "bill_of_lading_number"}
    inv = row.get("invoice_number")
    if inv:
        s = db.query(Shipment).filter(Shipment.invoice_number == inv).first()
        if s:
            return {"id": s.id, "shp_id": s.shp_id, "matched_by": "invoice_number"}
    cn = row.get("container_number")
    if cn:
        c = db.query(Container).filter(Container.container_number == cn).first()
        if c:
            s = db.query(Shipment).filter(Shipment.id == c.shipment_id).first()
            if s:
                return {"id": s.id, "shp_id": s.shp_id, "matched_by": "container_number"}
    return None


# =====================================================================
# Preview + Apply
# =====================================================================

def preview(db: Session, content: bytes) -> Dict[str, Any]:
    """Parse the file + classify each row as create/update/skip with reasons.
    Does NOT write anything."""
    rows, file_errors = parse_workbook(content)
    if file_errors:
        return {"file_errors": file_errors, "rows": [], "summary": {}}

    out_rows: List[Dict[str, Any]] = []
    counts = {"create": 0, "update": 0, "error": 0, "skip": 0}

    # Group multi-container rows by shipment_reference (or container_number+supplier)
    # so the user sees one suggested shipment per group.
    for r in rows:
        match = find_duplicate_match(db, r)
        if r.get("_errors"):
            action = "error"
        elif match:
            action = "update"
        else:
            action = "create"
        counts[action] += 1
        out_rows.append({
            **{k: v.isoformat() if isinstance(v, date) else v
               for k, v in r.items() if not k.startswith("_")},
            "_row": r.get("_row"),
            "_errors": r.get("_errors", []),
            "_match": match,
            "_action_default": action,
        })

    summary = {
        "total_rows": len(out_rows),
        **counts,
        "unique_suppliers": len({r.get("supplier_name") for r in out_rows if r.get("supplier_name")}),
        "unique_containers": len({r.get("container_number") for r in out_rows if r.get("container_number")}),
    }
    return {"file_errors": [], "rows": out_rows, "summary": summary}


def _row_to_shipment_kwargs(row: Dict[str, Any]) -> Dict[str, Any]:
    """Map our Excel keys → Shipment column kwargs."""
    return {
        "supplier":         row.get("supplier_name"),
        "category":         row.get("category"),
        "goods_description": row.get("notes"),
        "po_number":        row.get("purchase_order_number"),
        "invoice_number":   row.get("invoice_number"),
        "bol_number":       row.get("bill_of_lading_number"),
        "booking_number":   None,
        "customs_broker":   row.get("forwarder_name"),
        "origin_country":   row.get("origin_country"),
        "origin_port":      row.get("origin_port"),
        "shipping_channel": row.get("shipping_company"),
        "etd":              row.get("etd"),
        "eta_port":         row.get("eta_port"),
        "eta_warehouse":    row.get("eta_warehouse"),
        "actual_arrival_port": row.get("actual_arrival_port"),
        "actual_arrival_warehouse": row.get("actual_arrival_warehouse"),
        "stage_status":     row.get("shipment_status"),
        "notes":            row.get("notes"),
    }


def _row_to_container_kwargs(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "container_number": row.get("container_number"),
        "container_type":   row.get("container_type"),
        "boxes_total":      row.get("number_of_cartons"),
        "cbm":              row.get("cbm"),
        "gross_weight_kg":  row.get("gross_weight"),
        "eta_port":         row.get("eta_port"),
        "eta_warehouse":    row.get("eta_warehouse"),
        "actual_arrival_port": row.get("actual_arrival_port"),
        "actual_arrival_warehouse": row.get("actual_arrival_warehouse"),
        "category":         row.get("category"),
    }


def apply(db: Session, approved_rows: List[Dict[str, Any]],
          *, actor_name: str) -> Dict[str, Any]:
    """Apply the user-approved rows. Each row must include `_action` ∈
    {create, update, skip}. Rows with errors are silently skipped."""
    counts = {"created_shipments": 0, "updated_shipments": 0,
              "added_containers": 0, "updated_containers": 0,
              "skipped": 0}
    log_details: List[str] = []

    # Group rows by (shipment_reference OR matched id) so multi-container
    # rows roll into one shipment.
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for r in approved_rows:
        if r.get("_errors"): continue
        if r.get("_action") == "skip":
            counts["skipped"] += 1
            continue
        key = (str(r.get("shipment_reference") or "").strip()
               or f"_match_{(r.get('_match') or {}).get('id')}"
               or f"_row_{r.get('_row')}")
        groups.setdefault(key, []).append(r)

    for key, group in groups.items():
        head = group[0]
        action = head.get("_action") or head.get("_action_default") or "create"
        match = head.get("_match")

        # Coerce ISO date strings → date objects (frontend sends strings)
        for r in group:
            for fld in ("etd", "eta_port", "eta_warehouse",
                        "actual_arrival_port", "actual_arrival_warehouse"):
                if isinstance(r.get(fld), str) and r[fld]:
                    try: r[fld] = date.fromisoformat(r[fld])
                    except Exception: r[fld] = None

        ship_kwargs = _row_to_shipment_kwargs(head)

        if action == "update" and match:
            s = db.query(Shipment).filter(Shipment.id == match["id"]).first()
            if not s:
                counts["skipped"] += 1
                continue
            # Update only non-empty incoming values; respect manual_overrides.
            overrides = s.manual_overrides or {}
            updated_fields = []
            for col, val in ship_kwargs.items():
                if val in (None, ""): continue
                if col in overrides: continue   # manual override — never overwrite
                if getattr(s, col, None) != val:
                    setattr(s, col, val)
                    updated_fields.append(col)
            s.last_update_source = "excel_import"
            s.updated_by = actor_name
            if not s.data_source:
                s.data_source = "excel"
            counts["updated_shipments"] += 1
            event_service.log_event(
                db, entity_type="shipment", entity_id=s.id,
                action_type="excel_import_update",
                new_value=f"updated {len(updated_fields)} fields",
                changed_by=actor_name, source="excel_import",
                note=", ".join(updated_fields[:10]),
            )
        else:
            # Create
            shp_id = (head.get("shipment_reference") or "").strip() \
                     or shipment_service.next_shp_id(db)
            # Avoid duplicate shp_id collision
            if db.query(Shipment).filter(Shipment.shp_id == shp_id).first():
                shp_id = shipment_service.next_shp_id(db)
            s = Shipment(
                shp_id=shp_id,
                creation_source="excel_import",
                data_source="excel",
                is_test_data=False,
                last_update_source="excel_import",
                updated_by=actor_name,
                created_date=date.today(),
                **ship_kwargs,
            )
            db.add(s)
            db.flush()
            counts["created_shipments"] += 1
            event_service.log_event(
                db, entity_type="shipment", entity_id=s.id,
                action_type="excel_import_create",
                new_value=s.shp_id,
                changed_by=actor_name, source="excel_import",
            )

        # Containers: one per row, dedup by container_number
        for r in group:
            cn = (r.get("container_number") or "").strip()
            if not cn: continue
            existing = db.query(Container).filter(
                Container.container_number == cn,
            ).first()
            cont_kwargs = _row_to_container_kwargs(r)
            if existing:
                if existing.shipment_id != s.id:
                    log_details.append(
                        f"container {cn} כבר משויך ל-shipment#{existing.shipment_id} — דילגנו"
                    )
                    continue
                # Update fields where empty
                for col, val in cont_kwargs.items():
                    if val in (None, ""): continue
                    if getattr(existing, col, None) in (None, ""):
                        setattr(existing, col, val)
                counts["updated_containers"] += 1
            else:
                c = Container(
                    shipment_id=s.id,
                    updated_by=actor_name,
                    **cont_kwargs,
                )
                db.add(c)
                db.flush()
                counts["added_containers"] += 1

    db.commit()
    return {**counts, "details": log_details[:50]}
