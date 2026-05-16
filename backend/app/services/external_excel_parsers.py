"""ICL + Eli Line shipment-Excel parsers.

PURE FUNCTIONS — no DB access. Each parser takes raw bytes and returns
a structured preview list. Caller (the preview endpoint) joins this with
duplicate detection against the live shipments table.

DESIGN RULES (per the user spec):
  - NEVER invent missing data. Empty cell → null + needs_review marker.
  - Preserve every original raw value (container_raw, product_description_raw,
    container_quantity_raw) so the user can audit later.
  - Rounded values (e.g. 4.999 → 5) carry confidence='rounded_from_decimal'.
  - Dates: ICL is Excel serial number, Eli Line is dd.mm.yyyy text.
  - Excel sample-row markers (SAMPLE-*) are NOT relevant here — these are
    different formats from Royal Linen's own template.
  - Brand/category inference is SUGGESTION ONLY (separate fields).
"""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# =====================================================================
# Excel serial date conversion
# =====================================================================
# Excel epoch = 1900-01-00 with the famous 1900-02-29 leap-year bug.
# Practical formula for 1900-mode workbooks (the default since 2007):
#     date = datetime(1899, 12, 30) + timedelta(days=serial)
# That single-line formula correctly handles serials > 60.

EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_iso(serial) -> Optional[str]:
    """Convert an Excel serial number → ISO date string YYYY-MM-DD.
    Returns None if input isn't a parsable serial."""
    if serial is None or serial == "":
        return None
    if isinstance(serial, (datetime, date)):
        # openpyxl already parsed it
        d = serial.date() if isinstance(serial, datetime) else serial
        return d.isoformat()
    try:
        n = float(serial)
    except (TypeError, ValueError):
        return None
    if n < 1 or n > 2_000_000:
        return None
    try:
        d = (EXCEL_EPOCH + timedelta(days=n)).date()
        return d.isoformat()
    except Exception:
        return None


# =====================================================================
# Eli Line dd.mm.yyyy text date
# =====================================================================

_ELI_DATE_RX = re.compile(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\s*$")


def parse_eli_date(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        return d.isoformat()
    s = str(v).strip()
    if not s:
        return None
    m = _ELI_DATE_RX.match(s)
    if not m:
        return None
    d, mn, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return date(y, mn, d).isoformat()
    except ValueError:
        return None


# =====================================================================
# Container parsing for Eli Line column D ("סוג וכמות קונטיינרים")
# =====================================================================
#
# Examples observed in the user's spec:
#   "3X40HC"       → qty=3,  type="40HC"
#   "1X40HC"       → qty=1,  type="40HC"
#   "4X40HC+1X40"  → qty=5,  type="mixed 40HC/40"
#   "35-40 cbm"    → qty=null, type="LCL/CBM", cbm_raw="35-40 cbm",
#                    needs_review (range, not actual qty)

_CONTAINER_SEGMENT_RX = re.compile(r"(\d+)\s*[xXט×]\s*(\d{2,3})\s*([A-Z]{0,3})", re.IGNORECASE)
_CBM_RX = re.compile(r"(\d+(?:[.\-]\d+)?)\s*(?:-\s*\d+)?\s*cbm", re.IGNORECASE)


def parse_container_field(raw) -> Dict[str, Any]:
    """Parse Eli Line's container field.

    Returns:
      {
        container_raw: <original>,
        container_quantity: int|null,
        container_type: str|null,
        cbm_raw: str|null,
        needs_review: bool,
        review_reason: str|null,
      }
    """
    out = {
        "container_raw": None, "container_quantity": None,
        "container_type": None, "cbm_raw": None,
        "needs_review": False, "review_reason": None,
    }
    if raw is None:
        return out
    s = str(raw).strip()
    if not s:
        return out
    out["container_raw"] = s
    upper = s.upper().replace("×", "X")  # × → X

    # CBM variant — measured volume, not actual containers
    if "CBM" in upper:
        out["container_type"] = "LCL/CBM"
        out["cbm_raw"] = s
        out["needs_review"] = True
        out["review_reason"] = "Container field is CBM (volume), not container quantity"
        return out

    # NXTYPE [+ NXTYPE]+ pattern
    segments = _CONTAINER_SEGMENT_RX.findall(upper)
    if segments:
        total_qty = 0
        types = []
        for q, num, suffix in segments:
            total_qty += int(q)
            t = f"{num}{suffix}".upper().strip()
            if t and t not in types:
                types.append(t)
        out["container_quantity"] = total_qty
        if len(types) == 1:
            out["container_type"] = types[0]
        else:
            out["container_type"] = "mixed " + "/".join(types)
        return out

    # Fall-through: unknown shape — keep raw and flag review
    out["needs_review"] = True
    out["review_reason"] = f"Unrecognised container value: {s!r}"
    return out


# =====================================================================
# ICL container quantity (column L) — decimal that should round to int
# =====================================================================

def parse_icl_container_quantity(raw) -> Tuple[Optional[int], Optional[str], str]:
    """Returns (quantity, raw_str, confidence).
    confidence ∈ {"exact", "rounded_from_decimal", "zero", "unparseable"}

    Note: ICL stores container quantity as a high-precision decimal
    (e.g. 4.9994231423 actually means "5 containers"). Any non-integer
    input is therefore tagged `rounded_from_decimal` so the user knows
    the value didn't arrive as a clean integer.
    """
    if raw is None or raw == "":
        return None, None, "unparseable"
    if isinstance(raw, str) and not raw.strip():
        return None, None, "unparseable"
    raw_str = str(raw)
    try:
        n = float(raw)
    except (TypeError, ValueError):
        return None, raw_str, "unparseable"
    if n == 0:
        return 0, raw_str, "zero"
    rounded = round(n)
    # "exact" only if the input is a clean integer with no fractional part
    if n == int(n):
        return rounded, raw_str, "exact"
    return rounded, raw_str, "rounded_from_decimal"


# =====================================================================
# Brand / category inference (suggestion only — never overwrites)
# =====================================================================

INFERENCE_RULES: List[Dict[str, Any]] = [
    {"keywords": ["nandan", "מגבת", "מגבות", "חלוק מגבת"],
     "brand": None, "category": "Towels / Bath", "confidence": 0.85},
    {"keywords": ["puma", "ulac"],
     "brand": "Puma", "category": "Apparel / Footwear", "confidence": 0.85},
    {"keywords": ["טומי", "tommy"],
     "brand": "Tommy Hilfiger", "category": "Apparel", "confidence": 0.7},
    {"keywords": ["keds", "נעלי טבע", "e-go footwear", "e- go footwear", "ego footwear"],
     "brand": "Keds", "category": "Shoes", "confidence": 0.85},
    {"keywords": ["polder", "פולדר", "guangdong kenries", "kenries"],
     "brand": "Polder", "category": "Home / Laundry / Ironing", "confidence": 0.85},
    {"keywords": ["nautica", "נאוטיקה", "כריות"],
     "brand": "Nautica", "category": "Pillows / Bedding", "confidence": 0.7},
    {"keywords": ["lifetime", "life time", "צידניות", "צידנית"],
     "brand": "Lifetime", "category": "Coolers / Outdoor", "confidence": 0.8},
]


def infer_brand_category(*texts: Optional[str]) -> Dict[str, Any]:
    """Returns {inferred_brand, inferred_category, inference_confidence,
    matched_rule}. None values when no match."""
    blob = " ".join((t or "") for t in texts).lower()
    if not blob.strip():
        return {"inferred_brand": None, "inferred_category": None,
                "inference_confidence": 0.0, "matched_rule": None}
    best = None
    for r in INFERENCE_RULES:
        for kw in r["keywords"]:
            if kw.lower() in blob:
                if best is None or r["confidence"] > best["confidence"]:
                    best = r
                break
    if not best:
        return {"inferred_brand": None, "inferred_category": None,
                "inference_confidence": 0.0, "matched_rule": None}
    return {
        "inferred_brand": best["brand"],
        "inferred_category": best["category"],
        "inference_confidence": best["confidence"],
        "matched_rule": best["keywords"][0],
    }


# =====================================================================
# ICL parser
# =====================================================================

ICL_COLUMN_MAP = [
    # (column_letter, column_index, attr_name)
    ("A",  1, "icl_file_no"),
    ("B",  2, "sho_list"),
    ("C",  3, "customs_file_number"),
    ("D",  4, "purchase_order_number"),
    ("E",  5, "supplier_name"),
    ("F",  6, "origin_port"),
    ("G",  7, "destination_port"),
    ("H",  8, "incoterm"),
    ("I",  9, "carrier"),
    ("J", 10, "house_bill_of_lading_number"),
    ("K", 11, "master_bill_of_lading_number"),
    ("L", 12, "container_quantity_raw"),
    ("M", 13, "container_type"),
    ("N", 14, "etd_raw"),
    ("O", 15, "eta_port_raw"),
    ("P", 16, "product_description"),
]


def parse_icl(content: bytes, *, sheet_name: str, header_row: int,
              data_first_row: int, source_file_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    if sheet_name not in wb.sheetnames:
        # fall back to first sheet
        ws = wb.worksheets[0]
    else:
        ws = wb[sheet_name]

    rows_out: List[Dict[str, Any]] = []
    for r in range(data_first_row, ws.max_row + 1):
        # Empty row → skip
        if all(ws.cell(r, c[1]).value in (None, "") for c in ICL_COLUMN_MAP):
            continue

        raw: Dict[str, Any] = {}
        for _letter, idx, attr in ICL_COLUMN_MAP:
            v = ws.cell(r, idx).value
            if isinstance(v, str):
                v = v.strip() or None
            raw[attr] = v

        # Skip rows without ANY identifying field (all blanks of the few
        # required ones — defensive):
        if not (raw.get("icl_file_no") or raw.get("sho_list")
                or raw.get("supplier_name") or raw.get("product_description")):
            continue

        # ---- Field-level interpretation ----
        # shipment_reference: prefer ICL File No, else Sho List
        icl_file = raw.get("icl_file_no")
        sho_list = raw.get("sho_list")
        shipment_ref = icl_file if icl_file else sho_list
        ref_note = None
        if not icl_file and sho_list:
            ref_note = "Sho number only"

        # Container quantity
        cq, cq_raw, cq_conf = parse_icl_container_quantity(raw.get("container_quantity_raw"))

        # Dates (Excel serials)
        etd = excel_serial_to_iso(raw.get("etd_raw"))
        eta_port = excel_serial_to_iso(raw.get("eta_port_raw"))

        # Brand / category inference
        inf = infer_brand_category(
            raw.get("supplier_name"), raw.get("product_description"),
        )

        # Needs-review reasons
        review_reasons: List[str] = []
        if not raw.get("purchase_order_number"):
            review_reasons.append("PO is missing")
        if not raw.get("carrier"):
            review_reasons.append("Carrier is missing")
        if not raw.get("house_bill_of_lading_number") and not raw.get("master_bill_of_lading_number"):
            review_reasons.append("HBL and MBL both missing")
        elif not raw.get("house_bill_of_lading_number"):
            review_reasons.append("HBL is missing")
        if cq is None:
            review_reasons.append("Container quantity is missing or unparseable")
        elif cq == 0:
            review_reasons.append("Container quantity is zero")
        if not raw.get("container_type"):
            review_reasons.append("Container type is missing")
        if not etd:
            review_reasons.append("ETD is missing")
        if not eta_port:
            review_reasons.append("ETA (port) is missing")

        shipment_status = None
        if review_reasons and (not etd or not eta_port or cq in (None, 0)):
            shipment_status = "missing_shipping_details"
        elif eta_port:
            try:
                if date.fromisoformat(eta_port) < date.today():
                    shipment_status = "needs_update"  # ETA in the past
                else:
                    shipment_status = "in_transit"
            except Exception:
                pass

        rows_out.append({
            # Source provenance
            "source_provider": "ICL",
            "source_file_name": source_file_name,
            "source_sheet_name": sheet_name,
            "source_row_number": r,
            "data_source": "excel_import",
            "is_test_data": False,

            # Identifiers
            "shipment_reference": str(shipment_ref) if shipment_ref else None,
            "shipment_reference_note": ref_note,
            "external_file_number": str(icl_file) if icl_file else None,
            "sho_list": str(sho_list) if sho_list else None,
            "customs_file_number": str(raw.get("customs_file_number") or "") or None,
            "purchase_order_number": raw.get("purchase_order_number"),

            # Parties
            "supplier_name": raw.get("supplier_name"),

            # Geography
            "origin_port": raw.get("origin_port"),
            "destination_port": raw.get("destination_port"),
            "incoterm": raw.get("incoterm"),

            # Carrier + BL
            "carrier": raw.get("carrier"),
            "shipping_company": raw.get("carrier"),
            "house_bill_of_lading_number": raw.get("house_bill_of_lading_number"),
            "master_bill_of_lading_number": raw.get("master_bill_of_lading_number"),

            # Containers (quantity/type only — NOT actual numbers)
            "container_quantity_raw": cq_raw,
            "container_quantity": cq,
            "container_quantity_confidence": cq_conf,
            "container_type": raw.get("container_type"),
            "container_raw": raw.get("container_quantity_raw"),

            # Dates
            "etd": etd,
            "etd_raw": raw.get("etd_raw"),
            "eta_port": eta_port,
            "eta_port_raw": raw.get("eta_port_raw"),

            # Product
            "product_description": raw.get("product_description"),
            "product_description_raw": raw.get("product_description"),

            # Inference (suggestion only)
            **inf,

            # Status
            "shipment_status": shipment_status,
            "needs_review": bool(review_reasons),
            "review_reasons": review_reasons,
        })

    return rows_out


# =====================================================================
# Eli Line parser
# =====================================================================

ELI_COLUMN_MAP = [
    ("A", 1, "eta_port_raw"),
    ("B", 2, "vessel_name"),
    ("C", 3, "etd_raw"),
    ("D", 4, "container_field_raw"),
    ("E", 5, "product_description"),
    ("F", 6, "origin_port"),
    ("G", 7, "supplier_name"),
    ("H", 8, "external_job_number"),
    ("I", 9, "marks"),
]


def parse_eli_line(content: bytes, *, sheet_name: str,
                   data_first_row: int, source_file_name: str,
                   default_destination_port: str = "ASHDOD") -> List[Dict[str, Any]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    if sheet_name not in wb.sheetnames:
        ws = wb.worksheets[0]
    else:
        ws = wb[sheet_name]

    # Detect section headers like "ASHDOD", "HAIFA" — these set the
    # default destination_port for the rows below them, but are NOT
    # imported as shipments themselves.
    current_destination = default_destination_port

    rows_out: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        # Detect a "section header" row: row where only column A or B has
        # a port name and no JOB number / ETA in the expected columns.
        col_a = ws.cell(r, 1).value
        col_d = ws.cell(r, 4).value
        col_h = ws.cell(r, 8).value
        only_first_col_filled = (
            col_a and not col_d and not col_h
            and isinstance(col_a, str) and len(col_a.strip()) <= 30
        )
        if only_first_col_filled:
            # Treat as section header — update default destination + skip
            section = str(col_a).strip().upper()
            # Common Israeli destination ports / heuristic
            if section in ("ASHDOD", "HAIFA", "EILAT", "BEN GURION"):
                current_destination = section
            continue

        if r < data_first_row:
            continue

        # Empty row → skip
        if all(ws.cell(r, c[1]).value in (None, "") for c in ELI_COLUMN_MAP):
            continue

        raw: Dict[str, Any] = {}
        for _letter, idx, attr in ELI_COLUMN_MAP:
            v = ws.cell(r, idx).value
            if isinstance(v, str):
                v = v.strip() or None
            raw[attr] = v

        # Identifying field: external_job_number (column H)
        if not raw.get("external_job_number") and not raw.get("supplier_name"):
            continue

        # ---- Interpretation ----
        eta_port = parse_eli_date(raw.get("eta_port_raw"))
        etd = parse_eli_date(raw.get("etd_raw"))

        cont = parse_container_field(raw.get("container_field_raw"))

        # shipment_reference: JOB-XXXXX
        ext_job = raw.get("external_job_number")
        if ext_job is not None:
            try:
                ext_job_int = int(float(ext_job))
                ext_job_str = str(ext_job_int)
                shipment_ref = f"JOB-{ext_job_int}"
            except Exception:
                ext_job_str = str(ext_job)
                shipment_ref = f"JOB-{ext_job_str}"
        else:
            ext_job_str = None
            shipment_ref = None

        # marks
        marks = raw.get("marks")
        marks_str = str(marks) if marks is not None else None
        if marks_str and marks_str.replace(".", "").isdigit():
            try:
                marks_str = str(int(float(marks_str)))
            except Exception:
                pass

        # Brand / category inference
        inf = infer_brand_category(
            raw.get("supplier_name"), raw.get("product_description"),
        )

        # Needs-review reasons
        review_reasons: List[str] = list(
            [cont["review_reason"]] if cont["review_reason"] else []
        )
        if not eta_port:
            review_reasons.append("ETA is missing")
        if not etd:
            review_reasons.append("ETD is missing")
        if cont["container_quantity"] is None and cont["container_type"] != "LCL/CBM":
            review_reasons.append("Container quantity unknown")

        # Status
        shipment_status = None
        if not eta_port:
            shipment_status = "needs_review"
        else:
            try:
                if date.fromisoformat(eta_port) < date.today():
                    shipment_status = "needs_update"
                else:
                    shipment_status = "in_transit"
            except Exception:
                pass

        rows_out.append({
            "source_provider": "Eli Line",
            "source_file_name": source_file_name,
            "source_sheet_name": sheet_name,
            "source_row_number": r,
            "data_source": "excel_import",
            "is_test_data": False,

            "shipment_reference": shipment_ref,
            "external_job_number": ext_job_str,
            "marks": marks_str,

            "supplier_name": raw.get("supplier_name"),

            "origin_port": raw.get("origin_port"),
            "destination_port": current_destination,
            "vessel_name": raw.get("vessel_name"),

            "container_raw": cont["container_raw"],
            "container_quantity": cont["container_quantity"],
            "container_type": cont["container_type"],
            "cbm_raw": cont["cbm_raw"],

            "etd": etd,
            "etd_raw": raw.get("etd_raw"),
            "eta_port": eta_port,
            "eta_port_raw": raw.get("eta_port_raw"),

            "product_description": raw.get("product_description"),
            "product_description_raw": raw.get("product_description"),

            **inf,

            "shipment_status": shipment_status,
            "needs_review": bool(review_reasons),
            "review_reasons": review_reasons,
        })

    return rows_out
